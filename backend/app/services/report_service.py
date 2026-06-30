from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def build_sales_report_xlsx(rows, business_name: str, month_label: str, db_to_dash_status: dict) -> bytes:
    """rows: list of (Order, Customer) tuples. month_label: 'YYYY-MM'. Returns xlsx bytes."""
    wb = Workbook()

    # ── Sheet 1: Penjualan ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Penjualan"

    header_fill = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")

    # Row 1: title
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = f"Laporan Penjualan — {business_name} — {month_label}"
    title_cell.font = Font(bold=True, size=14)

    # Row 2: blank
    ws.append([])

    # Row 3: headers
    headers = ["Tanggal", "No. Pesanan", "Pelanggan", "Item", "Status", "Total"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # Data rows
    count_orders = 0
    sum_total = 0.0
    for order, customer in rows:
        count_orders += 1
        sum_total += float(order.total or 0)
        items_str = ", ".join(
            f"{it.get('name', '')} x{int(it.get('quantity') or it.get('qty') or 1)}"
            for it in (order.items or [])
        )
        pelanggan = customer.name or customer.phone_number
        status_display = db_to_dash_status.get(order.status, order.status)
        row_data = [
            order.created_at.strftime("%d/%m/%Y %H:%M"),
            f"#{order.order_seq:04d}",
            pelanggan,
            items_str,
            status_display,
            float(order.total or 0),
        ]
        ws.append(row_data)
        total_cell = ws.cell(row=ws.max_row, column=6)
        total_cell.number_format = "#,##0"

    # Blank row before summary
    ws.append([])

    # Summary block
    summary_row1 = ws.max_row + 1
    ws.append(["Total Pesanan", count_orders])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    ws.append(["Total Pendapatan", sum_total])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
    ws.cell(row=ws.max_row, column=2).number_format = "#,##0"

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14

    # ── Sheet 2: Per Produk ───────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Per Produk")

    ws2.append(["Produk", "Qty Terjual", "Total"])
    for col_idx in range(1, 4):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # Aggregate products
    product_map: dict[str, dict] = {}
    for order, _ in rows:
        for it in (order.items or []):
            raw_name = (it.get("name") or "").strip()
            if not raw_name:
                continue
            key = raw_name.lower()
            qty_each = int(it.get("quantity") or it.get("qty") or 1)
            price_each = float(it.get("price") or 0)
            if key not in product_map:
                product_map[key] = {"name": raw_name, "qty": 0, "total": 0.0}
            product_map[key]["qty"] += qty_each
            product_map[key]["total"] += qty_each * price_each

    sorted_products = sorted(product_map.values(), key=lambda x: x["qty"], reverse=True)
    for prod in sorted_products:
        ws2.append([prod["name"], prod["qty"], prod["total"]])
        ws2.cell(row=ws2.max_row, column=3).number_format = "#,##0"

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 12
    ws2.column_dimensions["C"].width = 14

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
